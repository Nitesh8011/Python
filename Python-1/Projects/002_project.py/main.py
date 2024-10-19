import openpyxl

data = openpyxl.load_workbook("/home/personal/Documents/Git/Python/Projects/002_project.py/inventory.xlsx")
sheet = data["Sheet1"]

supplier = {}
total_value_per_supplier = {}

for product_row in range(2, sheet.max_row + 1):
    supplier_name = sheet.cell(product_row, 4).value
    # print(supplier_name)

    # calculation number of products per supplier
    if supplier_name in supplier:
        current_num_product = supplier.get(supplier_name)
        # print(current_num_product)
        supplier[supplier_name] = current_num_product + 1
    else:
        supplier[supplier_name] = 1

# print(supplier)

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inve